from decimal import Decimal
from django.db import transaction
from django.core.exceptions import ValidationError
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.response import Response
from common.constants import VoucherStatus
from common.views import BaseViewSet
from journal_voucher.models import JournalVoucher, JournalEntry
from journal_voucher.serializers import (
    JournalVoucherSerializer,
    JournalVoucherListSerializer,
    ExcelUploadSerializer,
)
from journal_voucher.validators import (
    validate_debit_credit_balance,
    validate_period_open,
    validate_account_active,
)


class JournalVoucherViewSet(BaseViewSet):
    queryset = JournalVoucher.objects.all()
    search_fields = ['voucher_number', 'description', 'bl_number', 'invoice_number']

    def get_serializer_class(self):
        if self.action == 'list':
            return JournalVoucherListSerializer
        return JournalVoucherSerializer

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=['post'])
    def validate(self, request, pk=None):
        voucher = self.get_object()
        if voucher.status != VoucherStatus.DRAFT:
            return Response(
                {'error': f'Cannot validate a voucher with status {voucher.status}.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        entries = list(voucher.entries.all())
        if not entries:
            return Response(
                {'error': 'Voucher has no entries.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        errors = []
        try:
            validate_period_open(voucher.accounting_period)
        except ValidationError as e:
            errors.extend(e.messages if hasattr(e, 'messages') else [str(e)])
        try:
            validate_debit_credit_balance(entries)
        except ValidationError as e:
            errors.extend(e.messages if hasattr(e, 'messages') else [str(e)])
        for entry in entries:
            try:
                validate_account_active(entry.account)
            except ValidationError as e:
                errors.extend(e.messages if hasattr(e, 'messages') else [str(e)])
        if errors:
            return Response({'errors': errors}, status=status.HTTP_400_BAD_REQUEST)
        voucher.status = VoucherStatus.VALIDATED
        voucher.save(update_fields=['status'])
        serializer = self.get_serializer(voucher)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def void(self, request, pk=None):
        voucher = self.get_object()
        if voucher.status == VoucherStatus.VOID:
            return Response(
                {'error': 'Voucher is already void.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        voucher.status = VoucherStatus.VOID
        voucher.save(update_fields=['status'])
        serializer = self.get_serializer(voucher)
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def upload_excel(self, request):
        serializer = ExcelUploadSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        file = serializer.validated_data['file']

        try:
            from openpyxl import load_workbook
            wb = load_workbook(file, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            wb.close()
        except Exception as e:
            return Response(
                {'error': f'Failed to read Excel file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not rows:
            return Response(
                {'error': 'Excel file is empty.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        from chart_of_account.models import ChartOfAccount
        from accounting_period.models import AccountingPeriod

        results = {'created': 0, 'errors': []}

        vouchers_data = {}
        for row in rows:
            if not row[0]:
                continue
            key = str(row[0]).strip()
            if key not in vouchers_data:
                vouchers_data[key] = []
            vouchers_data[key].append(row)

        with transaction.atomic():
            for desc, lines in vouchers_data.items():
                try:
                    period_code = str(lines[0][6]).strip() if len(lines[0]) > 6 and lines[0][6] else None
                    raw_date = lines[0][5]
                    from datetime import datetime, date
                    if isinstance(raw_date, datetime):
                        tx_date = raw_date.date()
                    elif isinstance(raw_date, date):
                        tx_date = raw_date
                    elif raw_date:
                        try:
                            tx_date = datetime.strptime(str(raw_date).strip(), '%Y-%m-%d').date()
                        except (ValueError, TypeError):
                            tx_date = date.today()
                    else:
                        tx_date = date.today()

                    if not period_code:
                        raise ValidationError('Accounting Period Code is required in column G.')
                    period = AccountingPeriod.objects.get(code=period_code)

                    entries_data = []
                    for line in lines:
                        account_code = str(line[1]).strip() if line[1] else ''
                        if not account_code:
                            raise ValidationError('Account Code is required in column B.')
                        debit = Decimal(str(line[2])) if line[2] is not None else Decimal('0')
                        credit = Decimal(str(line[3])) if line[3] is not None else Decimal('0')
                        line_desc = str(line[4] or '').strip()

                        account = ChartOfAccount.objects.get(code=account_code)

                        if debit > 0 and credit > 0:
                            raise ValidationError(f'Entry for {account_code} has both debit and credit.')
                        if debit == 0 and credit == 0:
                            raise ValidationError(f'Entry for {account_code} has neither debit nor credit.')

                        entries_data.append({
                            'account': account,
                            'debit': debit,
                            'credit': credit,
                            'description': line_desc,
                        })

                    class EntryObj:
                        def __init__(self, d):
                            self.debit = d['debit']
                            self.credit = d['credit']

                    validate_debit_credit_balance([EntryObj(e) for e in entries_data])

                    voucher = JournalVoucher.objects.create(
                        accounting_period=period,
                        transaction_date=tx_date,
                        description=desc,
                        created_by=request.user if request.user.is_authenticated else None,
                    )

                    for i, ed in enumerate(entries_data):
                        JournalEntry.objects.create(
                            voucher=voucher,
                            account=ed['account'],
                            debit=ed['debit'],
                            credit=ed['credit'],
                            description=ed['description'],
                            line_order=i,
                        )

                    voucher.refresh_from_db()
                    all_entries = list(voucher.entries.all())
                    voucher.total_debit = sum(e.debit for e in all_entries)
                    voucher.total_credit = sum(e.credit for e in all_entries)
                    voucher.save(update_fields=['total_debit', 'total_credit'])

                    results['created'] += 1

                except Exception as e:
                    results['errors'].append(f'{desc}: {str(e)}')

        return Response(results)
